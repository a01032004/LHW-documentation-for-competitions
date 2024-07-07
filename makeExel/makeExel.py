from classes.Participant import Participant
from classes.Team import Team
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side # type: ignore
from openpyxl.styles import Alignment # type: ignore

'''book = openpyxl.open("first.xlsx", read_only=True)
sheet = book.active'''

# workbook = openpyxl.load_workbook("start.xlsx")
# print(sheet["i26"].value)
def make1exel(arr_teams, sort_arr_1, sort_arr_2, temp):
    """ToDO: _summary_

    Args:
        arr_teams (_type_): _description_
        sort_arr_1 (_type_): _description_
        sort_arr_2 (_type_): _description_
        temp (_type_): _description_
    """

    #создаём документ .xls
    wb = openpyxl.Workbook()
    
    #создаём страницу/лист "стартовый протокол"
    wb.create_sheet(title='Стартовый протокол', index=0)
    
    # получаем лист, с которым будем работать
    sheet = wb['Стартовый протокол']

    # создаем именованный стиль
    ns = NamedStyle(name='highlight')
    
    # создаем именованный стиль для текста
    ns.font = Font(bold=True, size=20, b=True)
    
    # создаем именованный стиль для границ табоицы
    border = Side(style='thick', color='000000')
    ns.border = Border(left=border, top=border, right=border, bottom=border)
    wb.add_named_style(ns)
    
    # начинаем формирование шапки стартового фацла

    # Объединяем ячейки A1:I1
    sheet.merge_cells('A1:I1')
    
    #Теперь объединённая ячейка стьала, как одна и в неё записываем текст
    sheet['A1'].value = 'СТАРТОВЫЙ ПРОТОКОЛ'
    
    #выравнивание текста
    sheet['A1'].alignment = Alignment(horizontal='center')
    
    #применнение стилей для текста
    sheet["A1"].font = Font(bold=True)

    sheet.merge_cells('A2:I2')
    sheet['A2'].value = 'СОРЕВНОВАНИЙ ПО ЛЫЖНЫМ ГОНКАМ'
    sheet['A2'].alignment = Alignment(horizontal='center')
    sheet["A2"].font = Font(bold=True)

    sheet.merge_cells('A3:I3')
    sheet['A3'].value = 'СПАРТАКИАДЫ А'
    sheet['A3'].alignment = Alignment(horizontal='center')
    sheet["A3"].font = Font(bold=True)
    sheet.merge_cells('A4:I4')
    
    # получаем группу соревнований
    type_competition = 1
    if (type_competition == 1):
        sheet['A4'].value = 'среди постоянного состава'
    if (type_competition == 2):
        sheet['A4'].value = 'среди переменного состава (мужчины)'
    if (type_competition == 3):
        sheet['A4'].value = 'среди переменного состава (женщины)'
    sheet['A4'].alignment = Alignment(horizontal='center')
    sheet["A4"].font = Font(bold=True)
    sheet.merge_cells('A6:D6')
    sheet['A6'].value = 'Дата проведения - 22.01.2024'
    sheet['A6'].alignment = Alignment(horizontal='left')
    sheet.merge_cells('E6:I6')
    sheet['E6'].value = 'Одинцовский парк культуры спорта и отдыха'
    sheet['E6'].alignment = Alignment(horizontal='left')
    sheet.merge_cells('A7:C7')
    sheet['A7'].value = 'Время старта - 14:50'
    sheet['A7'].alignment = Alignment(horizontal='left')
    sheet.merge_cells('A8:D8')
    if (type_competition == 1):
        sheet['A8'].value = 'Дистанция - 5 км (м), 3 км (ж)'
    if (type_competition == 2):
        sheet['A8'].value = 'Дистанция - 5 км'
    if (type_competition == 3):
        sheet['A8'].value = 'Дистанция - 3 км'
    sheet['A8'].alignment = Alignment(horizontal='left')
    sheet.merge_cells('E8:F8')
    sheet['E8'].value = 'Стиль - свободный'
    sheet['E8'].alignment = Alignment(horizontal='left')
    # Заполнение таблицы
    sheet['A10'].value = '№'
    sheet['B10'].value = 'Стартовый номер'
    sheet['C10'].value = 'Воинское звание'
    sheet['D10'].value = 'Ф.И.О.'
    sheet['E10'].value = 'Возраст'
    sheet['F10'].value = 'Команда'
    sheet['G10'].value = 'Время старта'
    sheet['H10'].value = 'Отметка о выходе со старта'
    sheet['A10'].alignment = Alignment(horizontal='center', vertical="center")
    sheet['B10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['C10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['D10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['E10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['F10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['G10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet['H10'].alignment = Alignment(
        horizontal='center', vertical="center", wrapText=True)
    sheet["A10"].font = Font(bold=True)
    sheet["B10"].font = Font(bold=True)
    sheet["C10"].font = Font(bold=True)
    sheet["D10"].font = Font(bold=True)
    sheet["E10"].font = Font(bold=True)
    sheet["F10"].font = Font(bold=True)
    sheet["G10"].font = Font(bold=True)
    sheet["H10"].font = Font(bold=True)
    sheet.row_dimensions[10].height = 58
    sheet.column_dimensions['H'].width = 10
    sheet.column_dimensions['D'].width = 38
    ns.font = Font(bold=False, size=20, b=False)
    
    #### pass    заполнение данных   ######
    count_peopple = 9
    count = 0
    for team in arr_teams:
        for partic in team.arr:
            sheet[f"A{count + 11}"] = count + 1
            sheet[f"B{count + 11}"] = partic.starting_number
            sheet[f"C{count + 11}"] = partic.rank
            sheet[f"D{count + 11}"] = partic.name
            sheet[f"E{count + 11}"] = partic.age
            sheet[f"F{count + 11}"] = partic.group
            sheet[f"G{count + 11}"] = partic.start_time
            sheet[f"H{count + 11}"] = partic.medical_allowance

            sheet[f"A{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"B{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"C{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"D{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"E{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"F{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"G{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"H{count + 11}"].alignment = Alignment(horizontal='center')
            count = count + 1
    
    # аполнение подписей
    sheet.merge_cells(f"A{count_peopple + 14}:C{count_peopple + 14}")
    sheet[f"A{count_peopple + 14}"].value = 'Главный судья'
    sheet[f"A{count_peopple + 14}"].alignment = Alignment(horizontal='left')
    sheet.merge_cells(f"E{count_peopple + 14}:F{count_peopple + 14}")
    sheet[f"E{count_peopple + 14}"].value = 'А.А.Ааааа'
    sheet.merge_cells(f"A{count_peopple + 15}:C{count_peopple + 15}")
    sheet[f"A{count_peopple + 15}"].value = 'Главный секретарь'
    sheet[f"A{count_peopple + 15}"].alignment = Alignment(horizontal='left')
    sheet.merge_cells(f"E{count_peopple + 15}:F{count_peopple + 15}")
    sheet[f"E{count_peopple + 15}"].value = 'Б.Б.Ббббб'
    sheet.merge_cells(f"A{count_peopple + 17}:B{count_peopple + 17}")
    sheet[f"A{count_peopple + 17}"].value = '22 января 2024г'
    
    # записываем файл и сохраняем
    wb.save('start.xlsx')






    ################################################     Итоговый     #############################################################
    #ele:
    if temp == 2:
        
        #создаём новый лист/старницу
        wn = wb.create_sheet('Итоговый протокол')
        
        # получаем лист, с которым будем работать
        sheet = wb['Итоговый протокол']
        
        # начинаем формирование шапки стартового протокола
        sheet.merge_cells('A1:L1')
        sheet['A1'].value = 'ИТОГОВЫЙ ПРОТОКОЛ'
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet["A1"].font = Font(bold=True)
        sheet.merge_cells('A2:L2')
        sheet['A2'].value = 'СОРЕВНОВАНИЙ ПО ЛЫЖНЫМ ГОНКАМ'
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet["A2"].font = Font(bold=True)
        sheet.merge_cells('A3:L3')
        sheet['A3'].value = 'СПАРТАКИАДЫ А'
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet["A3"].font = Font(bold=True)
        sheet.merge_cells('A4:L4')
        
        # получаем группу соревнований
        type_competition = 1
        if (type_competition == 1):
            sheet['A4'].value = 'среди постоянного состава'
        if (type_competition == 2):
            sheet['A4'].value = 'среди переменного состава (мужчины)'
        if (type_competition == 3):
            sheet['A4'].value = 'среди переменного состава (женщины)'
        sheet['A4'].alignment = Alignment(horizontal='center')
        sheet["A4"].font = Font(bold=True)
        sheet.merge_cells('A6:D6')
        sheet['A6'].value = 'Дата проведения - 22.01.2024'
        sheet['A6'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('E6:I6')
        sheet['E6'].value = 'Одинцовский парк культуры спорта и отдыха'
        sheet['E6'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('A7:C7')
        sheet['A7'].value = 'Время старта - 14:50'
        sheet['A7'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('A8:D8')
        if (type_competition == 1):
            sheet['A8'].value = 'Дистанция - 5 км (м), 3 км (ж)'
        if (type_competition == 2):
            sheet['A8'].value = 'Дистанция - 5 км'
        if (type_competition == 3):
            sheet['A8'].value = 'Дистанция - 3 км'
        sheet['A8'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('E8:F8')
        sheet['E8'].value = 'Стиль - свободный'
        sheet['E8'].alignment = Alignment(horizontal='left')
        
        # Заполнение таблицы
        sheet['A10'].value = '№'
        sheet['B10'].value = 'Стартовый номер'
        sheet['C10'].value = 'Воинское звание'
        sheet['D10'].value = 'Ф.И.О.'
        sheet['E10'].value = 'Возраст'
        sheet['F10'].value = 'Команда'
        sheet['G10'].value = 'Время старта'
        sheet['H10'].value = 'Время финиша'
        sheet['I10'].value = 'Чистое время'
        sheet['J10'].value = 'Коэфф.'
        sheet['K10'].value = 'Результат'
        sheet['L10'].value = 'Место'
        sheet['A10'].alignment = Alignment(horizontal='center', vertical="center")
        sheet['B10'].alignment = Alignment(horizontal='center', vertical="center")
        sheet['C10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['D10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['E10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['F10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['G10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['H10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['I10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['J10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['K10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['L10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet["A10"].font = Font(bold=True)
        sheet["B10"].font = Font(bold=True)
        sheet["C10"].font = Font(bold=True)
        sheet["D10"].font = Font(bold=True)
        sheet["E10"].font = Font(bold=True)
        sheet["F10"].font = Font(bold=True)
        sheet["G10"].font = Font(bold=True)
        sheet["H10"].font = Font(bold=True)
        sheet["I10"].font = Font(bold=True)
        sheet["J10"].font = Font(bold=True)
        sheet["K10"].font = Font(bold=True)
        sheet["L10"].font = Font(bold=True)

        #высота/ширина столбцов
        sheet.row_dimensions[10].height = 58
        sheet.column_dimensions['H'].width = 10
        sheet.column_dimensions['D'].width = 38
        
        #стили для текста
        ns.font = Font(bold=False, size=20, b=False)

        
        #### pass    заполнение данных   ######
        count_peopple = 9
        count = 0
        for partic in sort_arr_1:
            sheet[f"A{count + 11}"] = partic.id
            sheet[f"B{count + 11}"] = partic.starting_number
            sheet[f"C{count + 11}"] = partic.rank
            sheet[f"D{count + 11}"] = partic.name
            sheet[f"E{count + 11}"] = partic.age
            sheet[f"F{count + 11}"] = partic.group
            sheet[f"G{count + 11}"] = partic.start_time
            sheet[f"H{count + 11}"] = partic.finish_time
            sheet[f"I{count + 11}"] = partic.pure_time
            sheet[f"J{count + 11}"] = partic.factor
            sheet[f"K{count + 11}"] = partic.result_time
            sheet[f"L{count + 11}"] = partic.place

            sheet[f"A{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"B{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"C{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"D{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"E{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"F{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"G{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"H{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"G{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"H{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"G{count + 11}"].alignment = Alignment(horizontal='center')
            sheet[f"H{count + 11}"].alignment = Alignment(horizontal='center')
            count = count + 1

        # аполнение подписей
        sheet.merge_cells(f"A{count_peopple + 14}:C{count_peopple + 14}")
        sheet[f"A{count_peopple + 14}"].value = 'Главный судья'
        sheet[f"A{count_peopple + 14}"].alignment = Alignment(horizontal='left')
        sheet.merge_cells(f"E{count_peopple + 14}:F{count_peopple + 14}")
        sheet[f"E{count_peopple + 14}"].value = 'А.А.Ааааа'
        sheet.merge_cells(f"A{count_peopple + 15}:C{count_peopple + 15}")
        sheet[f"A{count_peopple + 15}"].value = 'Главный секретарь'
        sheet[f"A{count_peopple + 15}"].alignment = Alignment(horizontal='left')
        sheet.merge_cells(f"E{count_peopple + 15}:F{count_peopple + 15}")
        sheet[f"E{count_peopple + 15}"].value = 'Б.Б.Ббббб'
        sheet.merge_cells(f"A{count_peopple + 17}:B{count_peopple + 17}")
        sheet[f"A{count_peopple + 17}"].value = '22 января 2024г'
        wb.save('start.xlsx')


        ###########################################              Заключительная таблица                 ############################################

        #создаём новый лист/страницу
        wn = wb.create_sheet('Итоговая таблица')
        # получаем лист, с которым будем работать
        sheet = wb['Итоговая таблица']
        # начинаем формирование шапки стартового фацла
        sheet.merge_cells('A1:K1')
        sheet['A1'].value = 'ИТОГОВАЯ ТАБЛИЦА'
        sheet['A1'].alignment = Alignment(horizontal='center')
        sheet["A1"].font = Font(bold=True)
        sheet.merge_cells('A2:K2')
        sheet['A2'].value = 'СОРЕВНОВАНИЙ ПО ЛЫЖНЫМ ГОНКАМ'
        sheet['A2'].alignment = Alignment(horizontal='center')
        sheet["A2"].font = Font(bold=True)
        sheet.merge_cells('A3:K3')
        sheet['A3'].value = 'СПАРТАКИАДЫ А'
        sheet['A3'].alignment = Alignment(horizontal='center')
        sheet["A3"].font = Font(bold=True)
        sheet.merge_cells('A4:K4')
        # получаем группу соревнований
        type_competition = 1
        if (type_competition == 1):
            sheet['A4'].value = 'среди постоянного состава'
        if (type_competition == 2):
            sheet['A4'].value = 'среди переменного состава (мужчины)'
        if (type_competition == 3):
            sheet['A4'].value = 'среди переменного состава (женщины)'
        sheet['A4'].alignment = Alignment(horizontal='center')
        sheet["A4"].font = Font(bold=True)
        sheet.merge_cells('A6:D6')
        sheet['A6'].value = 'Дата проведения - 22.01.2024'
        sheet['A6'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('E6:I6')
        sheet['E6'].value = 'Одинцовский парк культуры спорта и отдыха'
        sheet['E6'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('A7:C7')
        sheet['A7'].value = 'Время старта - 14:50'
        sheet['A7'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('A8:D8')
        if (type_competition == 1):
            sheet['A8'].value = 'Дистанция - 5 км (м), 3 км (ж)'
        if (type_competition == 2):
            sheet['A8'].value = 'Дистанция - 5 км'
        if (type_competition == 3):
            sheet['A8'].value = 'Дистанция - 3 км'
        sheet['A8'].alignment = Alignment(horizontal='left')
        sheet.merge_cells('E8:F8')
        sheet['E8'].value = 'Стиль - свободный'
        sheet['E8'].alignment = Alignment(horizontal='left')
        # Заполнение таблицы
        sheet['A10'].value = ''
        sheet.merge_cells('B10:I10')
        sheet['B10'].value = 'Номера участников'
        sheet['J10'].value = 'Сумма 5 лучших'
        sheet['K10'].value = 'Место'
        sheet['A10'].alignment = Alignment(horizontal='center', vertical="center")
        sheet['B10'].alignment = Alignment(horizontal='center', vertical="center")
        sheet['C10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['D10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['E10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['F10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['G10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['H10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['I10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['J10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['K10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet['L10'].alignment = Alignment(
            horizontal='center', vertical="center", wrapText=True)
        sheet["A10"].font = Font(bold=True)
        sheet["B10"].font = Font(bold=True)
        sheet["J10"].font = Font(bold=True)
        sheet["K10"].font = Font(bold=True)
        #Ширина / высота столбцов и колонок
        sheet.row_dimensions[10].height = 48
        sheet.column_dimensions['A'].width = 18
        ns.font = Font(bold=False, size=20, b=False)

        
        #### pass    заполнение данных   ######
        count_peopple = 9
        pos = 0
        alphabet = ["B","C","D","E","F","G","H","I"]
        for team in sort_arr_2:
            count = 0
            sheet.merge_cells(f"A{11 + pos}:A{11 + pos + 1}")
            sheet[f"A{pos + 11}"] = team.team_name
            sheet[f"A{pos + 11}"].alignment = Alignment(vertical='center')
            sheet[f"A{pos + 11}"].alignment = Alignment(horizontal='center')
            for partic in team.arr:
                sheet[f"{alphabet[count]}{pos + 11}"] = partic.starting_number
                sheet[f"{alphabet[count]}{pos + 11}"].alignment = Alignment(horizontal='center')
                sheet[f"{alphabet[count]}{pos + 12}"] = partic.place
                sheet[f"{alphabet[count]}{pos + 12}"].alignment = Alignment(horizontal='center')
                count = count + 1
            sheet.merge_cells(f"J{11 + pos}:J{11 + pos + 1}")
            sheet[f"J{pos + 11}"] = team.team_points
            sheet[f"J{pos + 11}"].alignment = Alignment(vertical='center')
            sheet[f"J{pos + 11}"].alignment = Alignment(horizontal='center')
            sheet.merge_cells(f"K{11 + pos}:K{11 + pos + 1}")
            sheet[f"K{pos + 11}"] = team.place
            sheet[f"K{pos + 11}"].alignment = Alignment(vertical='center')
            sheet[f"K{pos + 11}"].alignment = Alignment(horizontal='center')
            pos = pos + 2


        # аполнение подписей
        sheet.merge_cells(f"A{count_peopple + 14}:C{count_peopple + 14}")
        sheet[f"A{count_peopple + 14}"].value = 'Главный судья'
        sheet[f"A{count_peopple + 14}"].alignment = Alignment(horizontal='left')
        sheet.merge_cells(f"E{count_peopple + 14}:F{count_peopple + 14}")
        sheet[f"E{count_peopple + 14}"].value = 'А.А.Ааааа'
        sheet.merge_cells(f"A{count_peopple + 15}:C{count_peopple + 15}")
        sheet[f"A{count_peopple + 15}"].value = 'Главный секретарь'
        sheet[f"A{count_peopple + 15}"].alignment = Alignment(horizontal='left')
        sheet.merge_cells(f"E{count_peopple + 15}:F{count_peopple + 15}")
        sheet[f"E{count_peopple + 15}"].value = 'Б.Б.Ббббб'
        sheet.merge_cells(f"A{count_peopple + 17}:B{count_peopple + 17}")
        sheet[f"A{count_peopple + 17}"].value = '22 января 2024г'
        #созраняем всё
        wb.save('start.xlsx')
