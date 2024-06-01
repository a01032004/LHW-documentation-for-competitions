from classes.Participant import Participant
from classes.Team import Team
import openpyxl
from openpyxl.styles import NamedStyle, Font, Border, Side
from openpyxl.styles import Alignment

'''book = openpyxl.open("first.xlsx", read_only=True)
sheet = book.active'''

#workbook = openpyxl.load_workbook("start.xlsx")
#print(sheet["i26"].value)

wb = openpyxl.Workbook()
wb.create_sheet(title = 'Первый лист', index = 0)
# получаем лист, с которым будем работать
sheet = wb['Первый лист']

# создаем именованный стиль
ns = NamedStyle(name='highlight')
ns.font = Font(bold=True, size=20)
border = Side(style='thick', color='000000')
ns.border = Border(left=border, top=border, right=border, bottom=border)
wb.add_named_style(ns)
#sheet['A1'].style = 'highlight'
#начинаем формирование шапки стартового фацла

sheet.merge_cells('A1:I1')
sheet['A1'].value = 'СТАРТОВЫЙ ПРОТОКОЛ'
sheet['A1'].alignment = Alignment(horizontal='center')

sheet.merge_cells('A2:I2')
sheet['A2'].value = 'СОРЕВНОВАНИЙ ПО ЛЫЖНЫМ ГОНКАМ'
sheet['A2'].alignment = Alignment(horizontal='center')

sheet.merge_cells('A3:I3')
sheet['A3'].value = 'СПАРТАКИАДЫ А'
sheet['A3'].alignment = Alignment(horizontal='center')

sheet.merge_cells('A4:I4')
#получаем группу соревнований
type_competition = 1
if (type_competition == 1):
    sheet['A4'].value = 'среди постоянного состава'
if (type_competition == 2):
    sheet['A4'].value = 'среди переменного состава (мужчины)'
if (type_competition == 3):
    sheet['A4'].value = 'среди переменного состава (женщины)'
sheet['A4'].alignment = Alignment(horizontal='center')


sheet.merge_cells('A6:D6')
sheet['A6'].value = 'Дата проведения - 22.01.2024'
sheet['A6'].alignment = Alignment(horizontal='left')

sheet.merge_cells('E6:I6')
sheet['E6'].value = 'Одинцовский парк культуры спорта и отдыха'
sheet['E6'].alignment = Alignment(horizontal='left')

sheet.merge_cells('A7:C7')
sheet['A7'].value = 'Время старта - 14:50'
sheet['A7'].alignment = Alignment(horizontal='left')

sheet.merge_cells('A8:D8')
if (type_competition == 1):
    sheet['A8'].value = 'Дистанция - 5 км (м), 3 км (ж)'
if (type_competition == 2):
    sheet['A8'].value = 'Дистанция - 5 км'
if (type_competition == 3):
    sheet['A8'].value = 'Дистанция - 3 км'
sheet['A8'].alignment = Alignment(horizontal='left')

sheet.merge_cells('E8:F8')
sheet['E8'].value = 'Стиль - свободный'
sheet['E8'].alignment = Alignment(horizontal='left')



#Заполнение таблицы
sheet['A10'].value = '№'
sheet['B10'].value = 'Стартовый номер'
sheet['C10'].value = 'Воинское звание'
sheet['D10'].value = 'Ф.И.О.'
sheet['E10'].value = 'Возраст'
sheet['F10'].value = 'Команда'
sheet['G10'].value = 'Время старта'
sheet['H10'].value = 'Отметка о выходе со старта'
sheet['A10'].alignment = Alignment(horizontal='center', vertical="center")
sheet['B10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['C10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['D10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['E10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['F10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['G10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)
sheet['H10'].alignment = Alignment(horizontal='center', vertical="center", wrapText=True)

sheet.row_dimensions[10].height = 52


####pass######

# записываем файл
wb.save('start.xlsx')





